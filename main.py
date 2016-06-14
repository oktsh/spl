import openpyxl
import csv
import xlwt
import datetime

# Открытие книги со списком наших серверов CAS
sl = openpyxl.load_workbook(filename='./CAS_server_list.xlsx')
sheet = sl['Sheet1']
servlist = [v[0].value.upper() for v in sheet.iter_rows('A1:A61')]  # создание list-списка серверов

# Открытие csv со списком серверов для патчинга
psl = open('ServerWindows.csv')
pslist = csv.reader(psl, dialect='excel')

psu = {}
i = 0

# Создание нового excel-файла
rw = xlwt.Workbook()
rws = rw.add_sheet('servers')

# Генерация заголовков рядов в excel-файле
rws.write(i, 0, 'Server name')
rws.write(i, 1, 'Patching date & time')
rws.write(i, 2, 'Patching week')
rws.write(i, 3, 'Reboot flag')

# Создание словаря с парами сервер:дата патчинга
for row in pslist:
    psu = dict(server=row[0].upper(), date=row[3], week=row[2], reboot=row[4])
    if psu['server'] in servlist:
        i += 1
        rws.write(i, 0, psu['server'])
        rws.write(i, 1, psu['date'])
        rws.write(i, 2, psu['week'])
        rws.write(i, 3, psu['reboot'])

# Сохранение файла с датой
date = str(datetime.date.today())
rw.save('Server patching list ' + date + '.xls')



